import openpyxl


def main():
    #This main function runs all the functions needed to read the scorecard excel file, update the good and bad holes from the data in the card, 
    #and write the new results into a new excel file known as ScoreProjectPython.xlsx
    path = "C:\\Users\\Kip\\Documents\\LAGCCScorecard\\LAGCC Score Project.xlsx"
    wb = open_wb(path)
    list = get_scores_to_par(wb)
    print(list)
    good_scores = solid_holes(list)
    print(good_scores)
    tough_scores = improvement_holes(list)
    print(tough_scores)
    good_holes = scores_to_holes(list, good_scores)
    print(good_holes)
    tough_holes = scores_to_holes(list, tough_scores)
    print(tough_holes)
    holes_to_xls(workbook=wb, holes=good_holes, solid=True)
    holes_to_xls(workbook=wb, holes=tough_holes, improve=True)


def get_scores_to_par(workbook):
    # This function reads the excel sheet workbook and creates a list of the average scores to par for all 18 holes.
    scores_to_par = []
    for value in workbook["Statistics"].iter_rows(
        min_row=3, max_row=3, min_col=2, max_col=10, values_only=True
    ):
        for x in range(9):
            scores_to_par.append(value[x])
    for value in workbook["Statistics"].iter_rows(
        min_row=3, max_row=3, min_col=12, max_col=20, values_only=True
    ):
        for x in range(9):
            scores_to_par.append(value[x])
    return scores_to_par


def solid_holes(list):
    # This function takes as input the list of average scores to par and creates a list of 3 or more solid holes, 
    # which does not include your best hole on the course.
    sort = sorted(list)
    if len(sort) < 4:
        raise ValueError("Insufficient Number of Holes")
    x = 0
    final = []
    for score in sort:
        if not (isinstance(score, float) or isinstance(score, int)):
            raise ValueError("Only Accepts Number Values")
        if x == 0:
            x += 1
            continue
        if x < 4:
            x += 1
            final.append(score)
        elif score in final:
            final.append(score)
    return final


def improvement_holes(list):
    # This function takes as input the list of average scores to par and creates a list of 3 or more holes which need improvement, 
    # which does not include your worst hole on the course.
    sort = sorted(list, reverse=True)
    if len(sort) < 4:
        raise ValueError("Insufficient Number of Holes")
    x = 0
    final = []
    for score in sort:
        if not (isinstance(score, float) or isinstance(score, int)):
            raise ValueError("Only Accepts Number Values")
        if x == 0:
            x += 1
            continue
        if x < 4:
            x += 1
            final.append(score)
        elif score in final:
            final.append(score)
    return final


def scores_to_holes(holes, scores):
    # This function takes as input the average score to par of every hole, and a smaller list of scores to par from other holes, 
    # returns as a list of strings the hole numbers of the scores of the smaller list.
    x = 1
    final = []
    for hole in holes:
        if hole in scores:
            final.append(f"Hole {x}")
        x += 1
    if len(final) != len(scores):
        raise ValueError("Different Data for Holes and Scores")
    return final


def holes_to_xls(workbook, holes, solid=False, improve=False):
    # This function takes as input the spreadsheet called workbook, the list of strings of hole numbers, a boolean value solid, and a boolean value improve.
    #A True value in solid or improve indicate which row of the excel sheet you would like to right the information in.
    if solid and improve:
        raise ValueError("Can only write to Solid or Improve at one time")
    elif not solid and not improve:
        raise ValueError("Must write to either Solid or Improve")
    if solid:
        ws = workbook["Statistics"]
        x = 0
        cells = [
            "B18",
            "C18",
            "D18",
            "E18",
            "F18",
            "G18",
            "H18",
            "I18",
            "J18",
            "K18",
            "L18",
            "M18",
            "N18",
            "018",
            "P18",
            "Q18",
            "R18",
            "S18",
        ]
        for hole in holes:
            ws[cells[x]] = hole
            x += 1
            workbook.save("ScoreProjectPython.xlsx")

    if improve:
        ws = workbook["Statistics"]
        x = 0
        cells = [
            "B17",
            "C17",
            "D17",
            "E17",
            "F17",
            "G17",
            "H17",
            "I17",
            "J17",
            "K17",
            "L17",
            "M17",
            "N17",
            "017",
            "P17",
            "Q17",
            "R17",
            "S17",
        ]
        for hole in holes:
            ws[cells[x]] = hole
            x += 1
            workbook.save("ScoreProjectPython.xlsx")

    return


def open_wb(path):
    # This function takes as input the directory path to the excel file we would like to pull data from, our scorecard, 
    # and returns the value as a variable which is usable by the library openpyxl.
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb


if __name__ == "__main__":
    main()
